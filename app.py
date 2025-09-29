import streamlit as st
import pandas as pd
from io import BytesIO
from PIL import Image as PILImage # Pillowをインポート
import requests
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from datetime import datetime
from openpyxl.utils.dataframe import dataframe_to_rows
import numpy as np

# ====== 固定設定 ======\
SHOP_ID = "lilirena"     # 楽天ショップID
IMG_MAX_H = 120          # 画像高さ(px)固定
IMG_COL_WIDTH = 18       # B列の幅(文字数)固定
# =====================

st.set_page_config(page_title="Order Maker", page_icon="🧾", layout="centered")

# --- NE風スタイル（青ボタン＆青タイトル） ---
st.markdown("""
<style>
:root { --ne-blue:#2a6df4; }
.block-container { max-width: 880px; }
.titlebar { font-size:22px; font-weight:800; display:flex; gap:10px; align-items:center; color:var(--ne-blue); }
.titlebar:before{content:"📄";}
.subtle{color:#667085; font-size:13px; margin-bottom:8px;}
.card{border:1px solid #e6e9ef; border-radius:14px; padding:22px; margin:14px 0; background:#fff;}
.stButton>button { background-color: var(--ne-blue); color: white; border-radius: 8px; border:none; padding:10px 20px; font-weight:600; transition:all 0.2s; }
.stButton>button:hover { background-color: #1e58e0; }
.stProgress > div > div > div > div { background-color: var(--ne-blue); }
.stDownloadButton > button { background-color: #10B981; } /* ダウンロードボタンは緑色に */
.stDownloadButton > button:hover { background-color: #049266; }
</style>
""", unsafe_allow_html=True)

# --- 関数群 ---

@st.cache_data
def load_data(uploaded_file):
    """アップロードされたファイルを読み込む (CSV/Excel)"""
    try:
        if uploaded_file.name.endswith('.csv'):
            return pd.read_csv(uploaded_file, encoding='cp932')
        elif uploaded_file.name.endswith(('.xlsx', '.xls')):
            return pd.read_excel(uploaded_file, engine='openpyxl')
        else:
            st.error("サポートされていないファイル形式です。CSVまたはExcelファイルをアップロードしてください。")
            return None
    except Exception as e:
        st.error(f"ファイルの読み込み中にエラーが発生しました: {e}")
        return None

def build_rakuten_url(sku):
    """SKUから楽天の商品URLを生成"""
    return f"https://item.rakuten.co.jp/{SHOP_ID}/{sku}/"

def download_image(url, referer=None):
    """画像をダウンロードし、BytesIO形式で返す"""
    if not url or "http" not in str(url):
        return None
    # User-Agentを追加し、ブロックされる可能性を下げる
    headers = {'Referer': referer, 
               'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'}
    try:
        response = requests.get(url, headers=headers, timeout=15)
        response.raise_for_status() # HTTPエラーチェック
        return BytesIO(response.content)
    except requests.exceptions.RequestException as e:
        return None

def process_image(img_io, sku): # SKUを引数に追加
    """
    画像をPILで処理し、Excel埋め込み用にファイルサイズを圧縮したBytesIOを返す (高速化のため)
    """
    if not img_io: return None
    try:
        img_io.seek(0)
        img = PILImage.open(img_io)
        
        # 縦横比を維持したまま高さをIMG_MAX_Hに合わせる
        ratio = IMG_MAX_H / img.height
        new_width = int(img.width * ratio)
        img_resized = img.resize((new_width, IMG_MAX_H), PILImage.Resampling.LANCZOS)
        
        # JPEG形式で圧縮してBytesIOに保存
        compressed_io = BytesIO()
        # quality=75で圧縮し、ファイルサイズを削減
        img_resized.save(compressed_io, format='JPEG', quality=75)
        compressed_io.seek(0)
        return compressed_io
    except Exception as e:
        # 画像ファイル形式が不正な場合の警告 (デバッグ用)
        st.warning(f"SKU: {sku} の画像処理中にエラーが発生しました。画像ファイル形式が不正な可能性があります。")
        return None

# --- メインロジック ---

st.markdown('<div class="titlebar">発注書作成アプリ</div>', unsafe_allow_html=True)
st.markdown('<div class="subtle">画像を自動取得し、Excel発注書に埋め込みます。</div>', unsafe_allow_html=True)

uploaded_file = st.file_uploader("ファイルをアップロードしてください", type=["csv", "xlsx", "xls"])

if uploaded_file is not None:
    df = load_data(uploaded_file)
    if df is not None:
        
        # 必須列の確認
        required_cols = ["sku", "数量", "原価", "特記事項", "商品名称", "商品URL", "画像URL(決定)"]
        if not all(col in df.columns for col in required_cols):
            missing_cols = [col for col in required_cols if col not in df.columns]
            # 必須カラム不足のエラーメッセージを強調
            st.error(f"🚨 **致命的なエラー：** 必須カラムが不足しています: **{', '.join(missing_cols)}**")
            st.info("アップロードするデータにこれらの列があるか、または列名が間違っていないか確認してください。")
            st.stop()
        
        total_rows = len(df)
        st.info(f"データ読込完了: 全 {total_rows} 行の処理を開始します。")

        # --- Excelの準備 ---
        wb = Workbook()
        ws = wb.active
        ws.title = "発注書"
        
        # ヘッダー行の挿入
        header_data = ["仕入先", "商品画像", "SKU/品番", "数量", "単価", 
                       "特記事項", "商品名", "商品URL", "備考1", "備考2", 
                       "備考3", "備考4", "小計", "合計", "発注日", "最終更新"]
        ws.append(header_data)
        
        # 列幅の設定
        for col_index, width in enumerate([15, IMG_COL_WIDTH, 15, 8, 10, 
                                            20, 30, 40, 10, 10, 
                                            10, 10, 10, 10, 15, 15]):
            ws.column_dimensions[ws.cell(row=1, column=col_index+1).column_letter].width = width

        # --- 処理ループ ---
        date_str = datetime.now().strftime('%Y/%m/%d')
        ok, fail = 0, 0
        
        # プログレスバーの表示（Excel埋め込み処理）
        prog2 = st.progress(0, text="Excel埋め込み 0%")
        
        for i, row in df.iterrows():
            # iは0から始まるインデックス、rowはシリーズ（行データ）
            qty = row.get("数量") if pd.notna(row.get("数量")) else 0
            genka = row.get("原価") if pd.notna(row.get("原価")) else 0
            gokei = qty * genka if pd.notna(qty) and pd.notna(genka) else None
            
            sku_val = row.get("sku") # SKUの値を取得
            img_url = row.get("画像URL(決定)") # 画像URLの値を取得

            excel_row = ["", "", sku_val, qty, genka,
                         row.get("特記事項"), row.get("商品名称"),
                         row.get("商品URL"), "", "", "", "",
                         genka, gokei, date_str, ""]
            ws.append(excel_row)

            r_i = ws.max_row
            referer = build_rakuten_url(sku_val) # SKUを使用

            # 1. 画像をダウンロード
            bin_io = download_image(img_url, referer=referer) 
            
            # 2. 画像を圧縮処理 (NEW!)
            processed_io = process_image(bin_io, sku_val) # SKUを渡す

            if processed_io:
                try:
                    # 3. 圧縮後のBytesIOを使用してExcel埋め込み
                    xlimg = XLImage(processed_io)
                    xlimg.anchor = f"B{r_i}"
                    ws.add_image(xlimg)
                    # 行の高さを設定 (画像サイズに合わせて)
                    ws.row_dimensions[r_i].height = int(IMG_MAX_H * 0.75)
                    ok += 1
                except Exception as e:
                    # 埋め込み失敗時のSKU表示 (デバッグ用)
                    st.warning(f"SKU: {sku_val} のExcel埋め込みに失敗しました。")
                    fail += 1
            else:
                fail += 1
                # ダウンロード失敗の原因を表示 (デバッグ用)
                if not img_url or "http" not in str(img_url):
                    st.warning(f"SKU: **{sku_val}** の**画像URLが空か無効**です。埋め込みをスキップしました。")
                else:
                    st.warning(f"SKU: **{sku_val}** (URL: {img_url[:50]}...) の**ダウンロードに失敗**しました。URLがブロックされている可能性があります。")

            # 進捗バーの更新（更新頻度を少し調整）
            prog2.progress(int((i+1)*100/total_rows), text=f"Excel埋め込み {i+1}/{total_rows} (画像成功: {ok} / 失敗: {fail})")

        prog2.progress(100, text="処理完了。発注書をダウンロードできます。")
        
        # --- ダウンロードボタン ---
        bio = BytesIO()
        try:
            wb.save(bio)
            bio.seek(0)
            filename = f"発注書_{datetime.now().strftime('%Y%m%d')}_画像埋め込み.xlsx"
            st.download_button("📥 発注書（画像埋め込み）をダウンロード",
                               data=bio,
                               file_name=filename,
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            st.success(f"処理完了: 画像埋め込み成功 {ok} 件 / 失敗 {fail} 件")
        except Exception as e:
            st.error(f"Excelファイルの保存中にエラーが発生しました: {e}")
